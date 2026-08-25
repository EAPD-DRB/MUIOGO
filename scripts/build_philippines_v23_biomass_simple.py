#!/usr/bin/env python3
"""Build the minimal, finite Philippines v23 biomass-supply candidate.

This is the promotion-path formulation.  It adds one crop-residue supply
technology and otherwise repairs existing source boundaries.  It performs no
model generation or optimization.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
import sys
from datetime import date
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
STORAGE = ROOT / "WebAPP" / "DataStorage"
SOURCE = STORAGE / "Philippines_v23"
DEFAULT_TARGET = STORAGE / ".Philippines_v23-biomass-supply-candidate-20260824"
API = ROOT / "API"
sys.path.insert(0, str(API))

BASE = "SC_0"
YEARS = tuple(str(year) for year in range(2020, 2054))
MIN_YEAR_SPLIT = 0.0004667
STANDARD_CAPACITY_TO_ACTIVITY = 31.536
RESIDUE_TECH = "PHL_PRO_SUP_CROP_RESIDUE"
RESIDUE_TECH_ID = "TEC_v23biors"

# Public 2020 crop observations and sourced residue physics.  Mt crop times
# t-residue/t-crop times GJ/t-residue is numerically PJ.
SUGARCANE_MT = 24.39894125
PALAY_MT = 14.57176518 + 4.72309034
BAGASSE_RPR = 0.29
BAGASSE_LHV = 16.45 * (1 - 0.50) - 2.443 * 0.50
RICE_HUSK_RPR = 0.225
RICE_HUSK_LHV = 2950 * 0.0041868
CANE_TRASH_RPR = 0.15
CANE_TRASH_LHV = 16.45 * (1 - 0.26) - 2.443 * 0.26
CANE_RECOVERY = 0.70 / 3.0
RESIDUE_COMPONENTS = {
    "bagasse": SUGARCANE_MT * BAGASSE_RPR * BAGASSE_LHV,
    "rice_husk": PALAY_MT * RICE_HUSK_RPR * RICE_HUSK_LHV * 0.95,
    "cane_trash": SUGARCANE_MT * CANE_TRASH_RPR * CANE_TRASH_LHV * CANE_RECOVERY,
}
RESIDUE_CAP_PJ = sum(RESIDUE_COMPONENTS.values())

# Conservative 2020 sustainable wood ceiling from public land areas and MAI.
WOOD_DENSITY_T_M3 = 0.725
WOOD_LHV = 15.23
FOREST_MHA = {"closed": 2.221173, "open": 4.693821, "mangrove": 0.311400}
FOREST_MAI = {"closed": 3.0, "open": 0.4, "mangrove": 0.0}
COCONUT_MHA = 3.65128876
ARABLE_MHA = 2.006 + 1.46544173 + 2.55378055 + 0.39908605 + 1.283887


def wood_pj(area_mha: float, mai_m3_ha: float) -> float:
    return area_mha * mai_m3_ha * WOOD_DENSITY_T_M3 * WOOD_LHV


def nonbinding_capacity_stock(annual_cap_pj: float) -> float:
    """Capacity stock whose rate envelope cannot bind before the annual cap."""
    return annual_cap_pj / (MIN_YEAR_SPLIT * STANDARD_CAPACITY_TO_ACTIVITY)


WOOD_COMPONENTS = {
    "forest": sum(wood_pj(FOREST_MHA[key], FOREST_MAI[key]) for key in FOREST_MHA),
    "coconut": wood_pj(COCONUT_MHA, 1.5),
    "arable": wood_pj(ARABLE_MHA, 1.0),
}
OTHER_BIOMASS_CAP_PJ = 0.836 * 41.868
GENERIC_CAP_PJ = sum(WOOD_COMPONENTS.values()) + OTHER_BIOMASS_CAP_PJ

# Constant-real 2020 supply costs from the disclosed build-ups requested for
# this correction.  These are supply costs, not tariffs or willingness to pay.
GENERIC_COST = 0.60 * (10.0 / 11.30) + 0.40 * ((500.0 / 0.375 / 49.62) / 11.30)
RESIDUE_COST = 0.576 * (10.0 / 7.00) + 0.304 * (20.50 / 12.38) + 0.120 * (20.50 / 11.61)

# Integrated legacy charcoal route: raw wood -> traditional kiln -> charcoal
# -> closed 20%-efficient stove stock.  No new charcoal technology is built.
CHARCOAL_ENERGY_YIELD = 0.175 * 27.82 / WOOD_LHV
CHARCOAL_ROUTE_IAR = 1.0 / (CHARCOAL_ENERGY_YIELD * 0.20)
KILN_CO2E_PER_USEFUL_PJ = (
    ((40.3 * 28 + 0.08 * 265) / 1000) * (1_000_000 / 27.82) / 1_000_000
) / 0.20
KILN_PM25_PER_USEFUL_PJ = (
    38.2 * (1 / CHARCOAL_ENERGY_YIELD) * 1_000_000 / WOOD_LHV / 1_000_000
) / 0.20
CHARCOAL_STOVE_PM25_PER_USEFUL_PJ = 0.2

# Public-data proxy for the inherited cooking-demand unit correction. DOE
# reports household biomass energy, while PSA reports households by primary
# cooking fuel; it does not report fuel-specific cooking energy. Therefore the
# fleet-weighted efficiency below is a disclosed proxy, not an exact conversion
# of the DOE biomass total into useful cooking service.
COOKING_SHARES = {"oil": 0.531, "electricity": 0.068, "charcoal": 0.072, "biomass": 0.325}
COOKING_EFFICIENCIES = {"oil": 0.60, "electricity": 0.70, "charcoal": 0.20, "biomass": 0.15}
COOKING_WEIGHTED_EFFICIENCY = (
    sum(COOKING_SHARES[k] * COOKING_EFFICIENCIES[k] for k in COOKING_SHARES)
    / sum(COOKING_SHARES.values())
)
COOKING_BIOMASS_INPUT_2020_PJ = 5.84188 * 41.868
COOKING_USEFUL_2020_PJ = COOKING_BIOMASS_INPUT_2020_PJ * COOKING_WEIGHTED_EFFICIENCY
COOKING_SCALE = COOKING_USEFUL_2020_PJ / 249.065


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=4) + "\n", encoding="utf-8")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def maps(gen):
    return (
        {row["Tech"]: row["TechId"] for row in gen["osy-tech"]},
        {row["Comm"]: row["CommId"] for row in gen["osy-comm"]},
        {row["Emis"]: row["EmisId"] for row in gen["osy-emis"]},
    )


def rows_for(table, parameter, tech_id, **coordinates):
    return [
        row for row in table[parameter][BASE]
        if row.get("TechId") == tech_id
        and all(row.get(key) == value for key, value in coordinates.items())
    ]


def one_row(table, parameter, tech_id, **coordinates):
    rows = rows_for(table, parameter, tech_id, **coordinates)
    if len(rows) != 1:
        raise AssertionError((parameter, tech_id, coordinates, len(rows)))
    return rows[0]


def set_years(row, value):
    for year in YEARS:
        row[year] = value


def inherit_rows(table, parameter, predicate):
    for scenario, rows in table[parameter].items():
        if scenario == BASE:
            continue
        for row in rows:
            if predicate(row):
                set_years(row, None)


def append_csv(path: Path, records: list[dict]) -> None:
    if not records:
        return
    with path.open(newline="", encoding="utf-8") as stream:
        reader = csv.DictReader(stream)
        fields = reader.fieldnames
        existing = list(reader)
    if fields is None:
        raise AssertionError(path)
    key = fields[0]
    existing = [row for row in existing if row[key] not in {record[key] for record in records}]
    with path.open("w", newline="", encoding="utf-8") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        writer.writerows(existing)
        writer.writerows({field: record.get(field, "") for field in fields} for record in records)


def mutate_structure(gen):
    tech_id, comm_id, emis_id = maps(gen)
    processor = next(row for row in gen["osy-tech"] if row["TechId"] == tech_id["PHL_PRO_PROC_BIOM"])
    processor.update({
        "Tech": "PHL_PRO_SUP_GENERIC_BIOMASS",
        "Desc": "Finite conservative 2020 domestic fuelwood and other-solid-biomass supply boundary; future use is endogenous.",
        "CapUnitId": "PJ/year", "ActUnitId": "PJ", "IAR": [],
        "OAR": [comm_id["PHL_PRO_BIOM"]], "EAR": [],
    })
    gen["osy-tech"].append({
        "TechId": RESIDUE_TECH_ID, "Tech": RESIDUE_TECH,
        "Desc": "Finite conservative 2020 crop-residue supply boundary; modes allocate one shared resource cap between legacy CHP and generic biomass.",
        "CapUnitId": "PJ/year", "ActUnitId": "PJ", "TG": [], "IAR": [],
        "OAR": [comm_id["PHL_PRO_BIOM_FIT_RESIDUE"], comm_id["PHL_PRO_BIOM"]],
        "INCR": [], "ITCR": [], "EAR": [],
    })
    ordinary = next(row for row in gen["osy-tech"] if row["TechId"] == tech_id["PHL_POW_CHP_BIOM_OLD"])
    ordinary["IAR"] = [
        comm_id["PHL_PRO_BIOM_FIT_RESIDUE"] if value == comm_id["PHL_PRO_BIOM"] else value
        for value in ordinary["IAR"]
    ]
    ordinary["Desc"] = ordinary["Desc"].rstrip(".") + "; uses the same finite crop-residue pool as the FIT slice."
    charcoal = next(row for row in gen["osy-tech"] if row["TechId"] == tech_id["PHL_HOU_COOK_COAL"])
    charcoal["Tech"] = "PHL_HOU_COOK_CHARCOAL_OLD"
    charcoal["Desc"] = "Closed legacy charcoal cooking route with kiln energy loss and upstream kiln emissions integrated into the route coefficients."
    charcoal["IAR"] = [comm_id["PHL_PRO_BIOM"]]
    charcoal["EAR"] = list(dict.fromkeys([*charcoal.get("EAR", []), emis_id["CO2e"], emis_id["PM2_5"]]))
    gen["osy-date"] = "2026-08-24"
    gen["osy-desc"] = (
        "Philippines v23 minimal biomass-supply repair. The founding undocumented biomass processor is reclassified as a finite sourced generic supply boundary; "
        "one finite residue boundary supplies both closed legacy CHP slices and may also supply generic users. Legacy charcoal losses and kiln emissions are integrated into its closed cooking route. "
        "No activity, dispatch, build or fuel share is forced. Conservative 2020 resource ceilings remain flat pending an endogenous crop/land formulation.\n\n"
        + gen["osy-desc"]
    )


def overlay_parameters(target: Path):
    gen = read_json(target / "genData.json")
    tech_id, comm_id, emis_id = maps(gen)
    generic = tech_id["PHL_PRO_SUP_GENERIC_BIOMASS"]
    residue = tech_id[RESIDUE_TECH]
    charcoal = tech_id["PHL_HOU_COOK_CHARCOAL_OLD"]

    ryt = read_json(target / "RYT.json")
    for tid, cap in ((generic, GENERIC_CAP_PJ), (residue, RESIDUE_CAP_PJ)):
        capacity_stock = nonbinding_capacity_stock(cap)
        for parameter, value in (("RC", capacity_stock), ("TAMaxC", capacity_stock), ("TAMaxCI", 0.0), ("TAU", cap), ("TAL", 0.0), ("AF", 1.0)):
            set_years(one_row(ryt, parameter, tid), value)
            inherit_rows(ryt, parameter, lambda row, tid=tid: row["TechId"] == tid)
    cooking = {
        tech_id[name] for name in (
            "PHL_HOU_COOK_OIL", "PHL_HOU_COOK_ELE", "PHL_HOU_COOK_NG",
            "PHL_HOU_COOK_CHARCOAL_OLD", "PHL_HOU_COOK_BIOM",
        )
    }
    parent_gen = read_json(SOURCE / "genData.json")
    parent_tid = {row["Tech"]: row["TechId"] for row in parent_gen["osy-tech"]}
    parent_ryt = read_json(SOURCE / "RYT.json")
    parent_by_parameter = {
        parameter: {row["TechId"]: row for row in parent_ryt[parameter][BASE]}
        for parameter in ("RC", "TAMaxCI")
    }
    for name in ("PHL_HOU_COOK_OIL", "PHL_HOU_COOK_ELE", "PHL_HOU_COOK_NG", "PHL_HOU_COOK_COAL", "PHL_HOU_COOK_BIOM"):
        candidate_name = "PHL_HOU_COOK_CHARCOAL_OLD" if name == "PHL_HOU_COOK_COAL" else name
        tid = tech_id[candidate_name]
        for parameter in ("RC", "TAMaxCI"):
            source = parent_by_parameter[parameter][parent_tid[name]]
            row = one_row(ryt, parameter, tid)
            for year in YEARS:
                row[year] = float(source[year]) * COOKING_SCALE
            inherit_rows(ryt, parameter, lambda item, tid=tid: item["TechId"] == tid)
    write_json(target / "RYT.json", ryt)

    rt = read_json(target / "RT.json")
    for parameter, value in (("CAU", STANDARD_CAPACITY_TO_ACTIVITY), ("OL", 1)):
        for scenario, rows in rt[parameter].items():
            for tid in (generic, residue):
                rows[0][tid] = value if scenario == BASE else None
    write_json(target / "RT.json", rt)

    rytm = read_json(target / "RYTM.json")
    for row in rows_for(rytm, "VC", generic):
        set_years(row, GENERIC_COST if row["MoId"] == 1 else 0.0)
    for row in rows_for(rytm, "VC", residue):
        set_years(row, RESIDUE_COST if row["MoId"] in (1, 2) else 0.0)
    inherit_rows(rytm, "VC", lambda row: row["TechId"] in {generic, residue})
    write_json(target / "RYTM.json", rytm)

    rytcm = read_json(target / "RYTCM.json")
    for row in rows_for(rytcm, "OAR", generic, CommId=comm_id["PHL_PRO_BIOM"]):
        set_years(row, 1.0 if row["MoId"] == 1 else 0.0)
    for row in rows_for(rytcm, "OAR", residue, CommId=comm_id["PHL_PRO_BIOM_FIT_RESIDUE"]):
        set_years(row, 1.0 if row["MoId"] == 1 else 0.0)
    for row in rows_for(rytcm, "OAR", residue, CommId=comm_id["PHL_PRO_BIOM"]):
        set_years(row, 1.0 if row["MoId"] == 2 else 0.0)
    for name in ("PHL_POW_CHP_BIOM_OLD", "PHL_POW_CHP_BIOM_FIT_OLD"):
        for row in rows_for(rytcm, "IAR", tech_id[name], CommId=comm_id["PHL_PRO_BIOM_FIT_RESIDUE"]):
            set_years(row, 4.000666667 if row["MoId"] == 1 else 0.0)
    for row in rows_for(rytcm, "IAR", charcoal, CommId=comm_id["PHL_PRO_BIOM"]):
        set_years(row, CHARCOAL_ROUTE_IAR if row["MoId"] == 1 else 0.0)
    affected = {generic, residue, charcoal, tech_id["PHL_POW_CHP_BIOM_OLD"], tech_id["PHL_POW_CHP_BIOM_FIT_OLD"]}
    inherit_rows(rytcm, "IAR", lambda row: row["TechId"] in affected)
    inherit_rows(rytcm, "OAR", lambda row: row["TechId"] in affected)
    write_json(target / "RYTCM.json", rytcm)

    rytts = read_json(target / "RYTTs.json")
    for tid in (generic, residue):
        for row in rows_for(rytts, "CF", tid):
            set_years(row, 1.0)
    inherit_rows(rytts, "CF", lambda row: row["TechId"] in {generic, residue})
    write_json(target / "RYTTs.json", rytts)

    ryc = read_json(target / "RYC.json")
    for parameter in ("SAD", "AAD"):
        for row in ryc[parameter][BASE]:
            if row["CommId"] == comm_id["PHL_HOU_COOK"]:
                for year in YEARS:
                    row[year] = float(row[year]) * COOKING_SCALE
        inherit_rows(ryc, parameter, lambda row: row["CommId"] == comm_id["PHL_HOU_COOK"])
    write_json(target / "RYC.json", ryc)

    rytem = read_json(target / "RYTEM.json")
    for emission, factor in (
        ("CO2e", KILN_CO2E_PER_USEFUL_PJ),
        ("PM2_5", CHARCOAL_STOVE_PM25_PER_USEFUL_PJ + KILN_PM25_PER_USEFUL_PJ),
    ):
        for row in rows_for(rytem, "EAR", charcoal, EmisId=emis_id[emission]):
            set_years(row, factor if row["MoId"] == 1 else 0.0)
    inherit_rows(rytem, "EAR", lambda row: row["TechId"] == charcoal)
    write_json(target / "RYTEM.json", rytem)


def write_workbook(ledger: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    workbook = Workbook()
    workbook.remove(workbook.active)
    for filename in ("SOURCES.csv", "CALCULATIONS.csv", "ASSUMPTIONS.csv", "MODEL_MAP.csv", "GAPS.csv", "CHANGES.csv"):
        sheet = workbook.create_sheet(filename[:-4][:31])
        with (ledger / filename).open(newline="", encoding="utf-8") as stream:
            for row in csv.reader(stream):
                sheet.append(row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
    path = ledger / "PHILIPPINES_V23_BIOMASS_SIMPLE_SCHEMA_LEDGER.xlsx"
    workbook.save(path)
    return path


def write_ledger(target: Path, source_hashes: dict[str, str]):
    ledger = target / "data_sources"
    shutil.copytree(SOURCE / "data_sources", ledger)
    sources = [
        {"source_id": "SRC_PHL_V23_BIOMASS_IRENA_SUPPLY", "provider": "International Renewable Energy Agency", "product": "Solid Biomass Supply for Heat and Power", "edition": "2019", "reference_period": "technology brief", "geography": "international", "variable": "at-source residue cost; haulage cost; dry lower heating values", "source_unit": "USD/t;USD/t-km;MJ/kg dry matter", "exact_locator": "feedstock cost and property tables", "url": "https://www.irena.org/publications/2019/Jan/Solid-biomass-supply-for-heat-and-power", "access_date": "2026-08-24", "license": "IRENA publication", "notes": "Supply-cost evidence; not a tariff or willingness-to-pay price."},
        {"source_id": "SRC_PHL_V23_BIOMASS_DOE_AWARDED_2025", "provider": "Philippine Department of Energy", "product": "Awarded Biomass Projects", "edition": "31 October 2025", "reference_period": "commercial-operation fleet", "geography": "Philippines", "variable": "solid-biomass capacity and named feedstock", "source_unit": "MW", "exact_locator": "commercial-operation entries", "url": "https://prod-cms.doe.gov.ph/documents/d/guest/a-1-biomass-pdf", "access_date": "2026-08-24", "license": "Philippine government publication", "notes": "Supports residue-cost weights; observed shares are benchmark-only."},
        {"source_id": "SRC_PHL_V23_BIOMASS_PSA_SUA_2020", "provider": "Philippine Statistics Authority", "product": "Supply Utilization Accounts of Selected Agricultural Commodities", "edition": "2018-2020", "reference_period": "2020", "geography": "Philippines", "variable": "sugarcane production", "source_unit": "metric tonnes", "exact_locator": "Table 6.4", "url": "https://psa.gov.ph/system/files/main-publication/SUA-2018-2020-ao13Nov_ONS-signed.pdf", "access_date": "2026-08-24", "license": "Philippine government publication", "notes": "2020 production 24,398,941 tonnes."},
        {"source_id": "SRC_PHL_V23_BIOMASS_FAO_MAI_SIMPLE", "provider": "Food and Agriculture Organization", "product": "Asia-Pacific Forestry Sector Outlook", "edition": "country tables", "reference_period": "long-run forestry characteristics", "geography": "Philippines", "variable": "mean annual wood increment by land class", "source_unit": "m3/ha/year", "exact_locator": "Table II.32", "url": "https://www.fao.org/4/w7714e/w7714e06.htm", "access_date": "2026-08-24", "license": "FAO publication", "notes": "Closed forest 3.0, open/low-productivity 0.4; coconut/arable values are disclosed conservative assumptions."},
        {"source_id": "SRC_PHL_V23_BIOMASS_DOE_KES_SIMPLE", "provider": "Philippine Department of Energy", "product": "2020 Key Energy Statistics", "edition": "2020", "reference_period": "2020", "geography": "Philippines", "variable": "other agricultural-waste biomass production", "source_unit": "ktoe", "exact_locator": "Biomass Production table", "url": "https://legacy.doe.gov.ph/sites/default/files/pdf/energy_statistics/2020_key_energy_statistics.pdf", "access_date": "2026-08-24", "license": "Philippine government publication", "notes": "836 ktoe converted at 41.868 PJ/Mtoe."},
        {"source_id": "SRC_PHL_V23_BIOMASS_NTA_FUELWOOD_SIMPLE", "provider": "National Tobacco Administration / Department of Agriculture", "product": "Fuelwood and Forestry Program", "edition": "public program document", "reference_period": "reported fuelwood price", "geography": "Philippines", "variable": "purchased fuelwood price", "source_unit": "PHP/m3", "exact_locator": "fuelwood price entry", "url": "https://nta.da.gov.ph/images/environment.pdf", "access_date": "2026-08-24", "license": "Philippine government publication", "notes": "PHP 500/m3 purchased-fuelwood input to the disclosed generic-basket cost."},
        {"source_id": "SRC_PHL_V23_BIOMASS_PSA_FOREST_SIMPLE", "provider": "Philippine Statistics Authority / NAMRIA", "product": "Land Asset Accounts of the Philippines", "edition": "2020 land cover", "reference_period": "2020", "geography": "Philippines", "variable": "closed, open and mangrove forest area", "source_unit": "hectares", "exact_locator": "2020 national land-cover values", "url": "https://psa.gov.ph/content/land-asset-accounts-philippines", "access_date": "2026-08-24", "license": "Philippine government publication", "notes": "Areas are retained in the local source snapshot."},
        {"source_id": "SRC_PHL_V23_BIOMASS_DOE_HOUSEHOLD", "provider": "Philippine Department of Energy", "product": "Compendium of Philippine Energy Statistics and Information", "edition": "2020", "reference_period": "2020", "geography": "Philippines", "variable": "household biomass final energy", "source_unit": "ktoe", "exact_locator": "Table 5", "url": "https://legacy.doe.gov.ph/energy-statistics/philippine-energy-balance", "access_date": "2026-08-24", "license": "Philippine government publication", "notes": "5,841.88 ktoe biomass."},
        {"source_id": "SRC_PHL_V23_BIOMASS_PSA_COOKING", "provider": "Philippine Statistics Authority", "product": "2020 Census of Population and Housing household characteristics", "edition": "2020", "reference_period": "2020", "geography": "Philippines", "variable": "primary household cooking-fuel shares", "source_unit": "percent of households", "exact_locator": "Figure 4", "url": "https://psa.gov.ph/content/household-characteristics-2020-census-population-and-housing", "access_date": "2026-08-24", "license": "Philippine government publication", "notes": "Used in a disclosed useful-cooking proxy and initial-stock rescaling; it is a household-share observation, not a fuel-energy share. Future activity shares remain endogenous."},
        {"source_id": "SRC_PHL_V23_BIOMASS_KILN_SIMPLE", "provider": "FAO; IPCC; Korean Society for Atmospheric Environment", "product": "Cebu woodfuel study; 2019 Refinement EFDB; commercial kiln field survey", "edition": "mixed", "reference_period": "technology factors", "geography": "Philippines/international", "variable": "charcoal yield; CH4/N2O; PM2.5", "source_unit": "fraction;g/kg;g/kg wood", "exact_locator": "15-20% Cebu yield; IPCC 1.B.1.c.i; 38.2 g/kg wood", "url": "https://www.fao.org/4/AD601E/ad601e00.pdf", "access_date": "2026-08-24", "license": "public scientific sources", "notes": "International kiln PM transfer is explicitly disclosed."},
    ]
    append_csv(ledger / "SOURCES.csv", sources)
    assumptions = [
        {"assumption_id": "ASM_PHL_V23_BIOMASS_SIMPLE_NONFORCING", "statement": "Supply shares, CHP dispatch and cooking route choices remain endogenous; only finite national resource ceilings are imposed.", "evidence_source_ids": "SRC_PHL_V23_BIOMASS_IRENA_SUPPLY;SRC_PHL_V23_BIOMASS_DOE_AWARDED_2025", "rationale": "Repository non-forcing calibration rule."},
        {"assumption_id": "ASM_PHL_V23_BIOMASS_SIMPLE_FLAT_CAP", "statement": "Public 2020 generic-biomass and crop-residue ceilings remain flat through 2053 pending an endogenous crop/forest formulation.", "central_value": json.dumps({"generic_pj": GENERIC_CAP_PJ, "residue_pj": RESIDUE_CAP_PJ}), "unit": "PJ/year", "evidence_source_ids": "SRC_PHL_V23_BIOMASS_PSA_SUA_2020;SRC_PSA_OPENSTAT_AGRICULTURE_2020;SRC_PHL_V23_BIOMASS_PSA_FOREST_SIMPLE;SRC_PHL_V23_BIOMASS_FAO_MAI_SIMPLE;SRC_PHL_V23_BIOMASS_DOE_KES_SIMPLE", "rationale": "A no-growth ceiling is conservative and avoids unsupported long-run resource growth; it is a physical upper bound, not an activity target."},
        {"assumption_id": "ASM_PHL_V23_BIOMASS_SIMPLE_COSTS", "statement": "The 2020 supply-cost build-ups remain constant in real model currency.", "central_value": json.dumps({"generic": GENERIC_COST, "residue": RESIDUE_COST}), "unit": "MUSD/PJ", "evidence_source_ids": "SRC_PHL_V23_BIOMASS_IRENA_SUPPLY;SRC_PHL_V23_BIOMASS_DOE_AWARDED_2025;SRC_PHL_V23_BIOMASS_NTA_FUELWOOD_SIMPLE;SRC_PSA_SSAF_2022", "rationale": "No defensible Philippine real-cost forecast was found."},
        {"assumption_id": "ASM_PHL_V23_BIOMASS_SIMPLE_CHARCOAL", "statement": "Traditional kiln loss and upstream emissions are integrated into the closed legacy charcoal cooking route instead of represented by a separate kiln technology.", "central_value": CHARCOAL_ENERGY_YIELD, "unit": "PJ charcoal/PJ raw wood", "evidence_source_ids": "SRC_PHL_V23_BIOMASS_KILN_SIMPLE", "rationale": "Preserve physical loss and emissions with minimal matrix structure."},
        {"assumption_id": "ASM_PHL_V23_BIOMASS_SIMPLE_COOKING_PROXY", "statement": "In the absence of fuel-specific 2020 cooking-only energy, normalized primary-cooking-fuel household shares weight route efficiencies and are applied to DOE total household biomass energy to proxy useful cooking service.", "central_value": COOKING_WEIGHTED_EFFICIENCY, "unit": "PJ useful/PJ final energy", "evidence_source_ids": "SRC_PHL_V23_BIOMASS_DOE_HOUSEHOLD;SRC_PHL_V23_BIOMASS_PSA_COOKING", "rationale": "Correct the inherited final-energy/useful-service unit mismatch without forcing route activity; the proxy and its validation limitation are explicit."},
    ]
    append_csv(ledger / "ASSUMPTIONS.csv", assumptions)
    calculations = [
        {"calculation_id": "CALC_PHL_V23_BIOMASS_SIMPLE_RESIDUE_CAP", "formula": "sum(crop Mt * residue/product * as-received LHV * recoverable fraction)", "source_ids": "SRC_PHL_V23_BIOMASS_PSA_SUA_2020;SRC_PSA_OPENSTAT_AGRICULTURE_2020;SRC_PHL_V23_BIOMASS_IRENA_SUPPLY", "assumption_ids": "ASM_PHL_V23_BIOMASS_SIMPLE_FLAT_CAP", "input_values": json.dumps(RESIDUE_COMPONENTS), "input_units": "PJ/year by residue", "output_value": RESIDUE_CAP_PJ, "output_unit": "PJ/year", "script_path": "scripts/build_philippines_v23_biomass_simple.py", "script_version": "v1"},
        {"calculation_id": "CALC_PHL_V23_BIOMASS_SIMPLE_GENERIC_CAP", "formula": "sum(public land area * sustainable MAI * 0.725 t/m3 * 15.23 GJ/t) + 0.836 Mtoe * 41.868", "source_ids": "SRC_PHL_V23_BIOMASS_FAO_MAI_SIMPLE;SRC_PHL_V23_BIOMASS_PSA_FOREST_SIMPLE;SRC_PSA_OPENSTAT_AGRICULTURE_2020;SRC_PHL_V23_BIOMASS_DOE_KES_SIMPLE", "assumption_ids": "ASM_PHL_V23_BIOMASS_SIMPLE_FLAT_CAP", "input_values": json.dumps({**WOOD_COMPONENTS, "other": OTHER_BIOMASS_CAP_PJ}), "input_units": "PJ/year", "output_value": GENERIC_CAP_PJ, "output_unit": "PJ/year", "script_path": "scripts/build_philippines_v23_biomass_simple.py", "script_version": "v1"},
        {"calculation_id": "CALC_PHL_V23_BIOMASS_SIMPLE_COSTS", "formula": "generic=0.6*(10/11.30)+0.4*((500/0.375/49.62)/11.30); residue=0.576*(10/7)+0.304*(20.5/12.38)+0.12*(20.5/11.61)", "source_ids": "SRC_PHL_V23_BIOMASS_IRENA_SUPPLY;SRC_PHL_V23_BIOMASS_DOE_AWARDED_2025;SRC_PHL_V23_BIOMASS_NTA_FUELWOOD_SIMPLE;SRC_PSA_SSAF_2022", "assumption_ids": "ASM_PHL_V23_BIOMASS_SIMPLE_COSTS", "input_values": "source values and disclosed weights", "input_units": "USD/t;GJ/t;fraction", "output_value": json.dumps({"generic": GENERIC_COST, "residue": RESIDUE_COST}), "output_unit": "MUSD/PJ", "script_path": "scripts/build_philippines_v23_biomass_simple.py", "script_version": "v1"},
        {"calculation_id": "CALC_PHL_V23_BIOMASS_SIMPLE_CHARCOAL", "formula": "route IAR=1/(0.175*27.82/15.23*0.20); kiln emissions divided by 0.20 stove efficiency", "source_ids": "SRC_PHL_V23_BIOMASS_KILN_SIMPLE", "assumption_ids": "ASM_PHL_V23_BIOMASS_SIMPLE_CHARCOAL", "input_values": f"yield={CHARCOAL_ENERGY_YIELD};IAR={CHARCOAL_ROUTE_IAR}", "input_units": "fraction;PJ/PJ useful", "output_value": json.dumps({"iar": CHARCOAL_ROUTE_IAR, "co2e": KILN_CO2E_PER_USEFUL_PJ, "pm25_total": CHARCOAL_STOVE_PM25_PER_USEFUL_PJ + KILN_PM25_PER_USEFUL_PJ}), "output_unit": "PJ/PJ;Mt/PJ;kt/PJ", "script_path": "scripts/build_philippines_v23_biomass_simple.py", "script_version": "v1"},
        {"calculation_id": "CALC_PHL_V23_BIOMASS_SIMPLE_COOKING", "formula": "proxy useful cooking = DOE household biomass final energy * normalized PSA cooking-household-share-weighted route efficiency; scale = proxy useful cooking / inherited 249.065 PJ", "source_ids": "SRC_PHL_V23_BIOMASS_DOE_HOUSEHOLD;SRC_PHL_V23_BIOMASS_PSA_COOKING", "assumption_ids": "ASM_PHL_V23_BIOMASS_SIMPLE_COOKING_PROXY", "input_values": json.dumps({"biomass_mtoe": 5.84188, "pj_per_mtoe": 41.868, "normalized_household_share_weighted_efficiency": COOKING_WEIGHTED_EFFICIENCY, "inherited_pj": 249.065}), "input_units": "Mtoe;PJ/Mtoe;fraction;PJ", "output_value": json.dumps({"biomass_input_2020_pj": COOKING_BIOMASS_INPUT_2020_PJ, "proxy_useful_2020_pj": COOKING_USEFUL_2020_PJ, "scale": COOKING_SCALE}), "output_unit": "PJ;PJ;fraction", "script_path": "scripts/build_philippines_v23_biomass_simple.py", "script_version": "v4"},
        {"calculation_id": "CALC_PHL_V23_BIOMASS_SIMPLE_CAPACITY", "formula": "CAU=31.536; RC=TAMaxC=annual_cap/(min(YearSplit)*CAU); TAU=annual_cap; TAMaxCI=0", "assumption_ids": "ASM_PHL_V23_BIOMASS_SIMPLE_NONFORCING", "input_values": json.dumps({"min_year_split": MIN_YEAR_SPLIT, "cau": STANDARD_CAPACITY_TO_ACTIVITY, "generic_rc": nonbinding_capacity_stock(GENERIC_CAP_PJ), "residue_rc": nonbinding_capacity_stock(RESIDUE_CAP_PJ)}), "input_units": "fraction;activity/capacity;capacity", "output_value": json.dumps({"generic_rate_envelope": GENERIC_CAP_PJ / MIN_YEAR_SPLIT, "residue_rate_envelope": RESIDUE_CAP_PJ / MIN_YEAR_SPLIT}), "output_unit": "PJ/year rate", "script_path": "scripts/build_philippines_v23_biomass_simple.py", "script_version": "v2"},
    ]
    append_csv(ledger / "CALCULATIONS.csv", calculations)
    model_maps = [
        {"map_id": "MAP_PHL_V23_BIOMASS_SIMPLE_GENERIC", "model_file": "genData.json;RT.json;RYT.json;RYTM.json;RYTCM.json;RYTTs.json", "parameter": "technology classification;RC;TAMaxC;TAMaxCI;TAU;CAU;CF;VC;OAR", "entity": "PHL_PRO_SUP_GENERIC_BIOMASS", "mode": "1", "scenario": "SC_0 with inheritance", "years": "2020-2053", "value_or_expression": "CALC_PHL_V23_BIOMASS_SIMPLE_GENERIC_CAP;CALC_PHL_V23_BIOMASS_SIMPLE_COSTS;CALC_PHL_V23_BIOMASS_SIMPLE_CAPACITY", "model_unit": "PJ/year;MUSD/PJ", "evidence_ids": "CALC_PHL_V23_BIOMASS_SIMPLE_GENERIC_CAP;CALC_PHL_V23_BIOMASS_SIMPLE_COSTS;CALC_PHL_V23_BIOMASS_SIMPLE_CAPACITY", "evidence_type": "finite domestic resource boundary"},
        {"map_id": "MAP_PHL_V23_BIOMASS_SIMPLE_RESIDUE", "model_file": "genData.json;RT.json;RYT.json;RYTM.json;RYTCM.json;RYTTs.json", "parameter": "new technology;RC;TAMaxC;TAMaxCI;TAU;CAU;CF;VC;OAR", "entity": RESIDUE_TECH, "mode": "1 CHP pool;2 generic biomass", "scenario": "SC_0 with inheritance", "years": "2020-2053", "value_or_expression": "one shared finite residue cap allocated endogenously between outputs", "model_unit": "PJ/year;MUSD/PJ", "evidence_ids": "CALC_PHL_V23_BIOMASS_SIMPLE_RESIDUE_CAP;CALC_PHL_V23_BIOMASS_SIMPLE_COSTS;CALC_PHL_V23_BIOMASS_SIMPLE_CAPACITY", "evidence_type": "finite domestic resource boundary"},
        {"map_id": "MAP_PHL_V23_BIOMASS_SIMPLE_CHP", "model_file": "genData.json;RYTCM.json", "parameter": "IAR membership;IAR", "entity": "PHL_POW_CHP_BIOM_OLD;PHL_POW_CHP_BIOM_FIT_OLD", "mode": "1", "scenario": "SC_0 with inheritance", "years": "2020-2053", "value_or_expression": "both consume PHL_PRO_BIOM_FIT_RESIDUE at 4.000666667", "model_unit": "PJ/PJ", "evidence_ids": "ASM_PHL_V23_BIOMASS_SIMPLE_NONFORCING", "evidence_type": "common physical feedstock pool"},
        {"map_id": "MAP_PHL_V23_BIOMASS_SIMPLE_CHARCOAL", "model_file": "genData.json;RYTCM.json;RYTEM.json", "parameter": "IAR;EAR", "entity": "PHL_HOU_COOK_CHARCOAL_OLD", "mode": "1", "scenario": "SC_0 with inheritance", "years": "2020-2053", "value_or_expression": "CALC_PHL_V23_BIOMASS_SIMPLE_CHARCOAL", "model_unit": "PJ/PJ;Mt/PJ;kt/PJ", "evidence_ids": "CALC_PHL_V23_BIOMASS_SIMPLE_CHARCOAL", "evidence_type": "integrated physical conversion and emissions"},
        {"map_id": "MAP_PHL_V23_BIOMASS_SIMPLE_COOKING", "model_file": "RYC.json;RYT.json", "parameter": "SAD;AAD;RC;TAMaxCI", "entity": "PHL_HOU_COOK;PHL_HOU_COOK_*", "scenario": "SC_0 with inheritance", "years": "2020-2053", "value_or_expression": "inherited values multiplied by CALC_PHL_V23_BIOMASS_SIMPLE_COOKING scale", "model_unit": "PJ useful/year;capacity", "evidence_ids": "CALC_PHL_V23_BIOMASS_SIMPLE_COOKING", "evidence_type": "final-demand unit correction and consistent stock rescaling"},
    ]
    append_csv(ledger / "MODEL_MAP.csv", model_maps)
    append_csv(ledger / "GAPS.csv", [
        {"item": "Endogenous crop/forest biomass availability", "why_absent": "The detailed crop-gate and land-coproduct formulation caused an excessive CBC runtime increase; the minimal formulation uses conservative flat 2020 caps.", "upgrade_source": "Reintroduce crop- and land-responsive supplies only with a numerically sparse formulation and a dedicated scenario-aware runtime A/B.", "priority": "high", "notes": "Until upgraded, changes in modeled crop or forest area do not alter biomass ceilings."},
        {"item": "Separate generated, collected and uncollected biomass reporting", "why_absent": "The minimal boundary reports delivered supply by destination but not physical residue generation or unused potential.", "upgrade_source": "A sparse physical residue account with soil-retention, open-burning and collection emissions.", "priority": "medium", "notes": "This detail is not required to correct the price, shared-CHP feedstock or cooking defect."},
        {"item": "Time-varying domestic biomass resource outlook", "why_absent": "No defensible national 2021-2053 crop-residue and sustainable-wood projection was identified.", "upgrade_source": "Official long-run crop, forest-cover, sustainable-harvest and residue-recovery scenarios.", "priority": "high", "notes": "Flat 2020 caps are conservative physical ceilings, not fixed activity."},
        {"item": "Fuel-specific household cooking-energy calibration", "why_absent": "DOE reports total household fuel use and PSA reports primary cooking fuel by household, but the available tables do not provide 2020 fuel-specific cooking-only energy. The promoted 105.436 PJ useful-service value is therefore a transparent proxy, not an exact end-use balance.", "upgrade_source": "A nationally representative household energy survey reporting cooking-only quantities by fuel and device, with matching stove efficiencies.", "priority": "high", "notes": "Use the observation to recalibrate final demand and initial stocks; do not force fuel shares or activity. The 2020 cooking PM2.5 benchmark is not improved by the present proxy."},
    ])
    append_csv(ledger / "CHANGES.csv", [{
        "change_id": "CHG_PHL_V23_BIOMASS_SIMPLE_20260824", "date": str(date(2026, 8, 24)), "class": "B",
        "description": "Minimal finite biomass-supply, shared legacy-CHP feedstock, useful-cooking and integrated-charcoal correction.",
        "model_objects": "genData.json;RT.json;RYT.json;RYTM.json;RYTCM.json;RYTTs.json;RYC.json;RYTEM.json;RYTCn.json",
        "evidence_path": "calculation_notes/MODEL_FIXES_BIOMASS_SIMPLE_V23_2026-08-24.md",
        "map_rows_affected": ";".join(row["map_id"] for row in model_maps),
        "resolve_status": "candidate_pending_zero_solve_gates_matrix_and_BASE_runtime_gate", "author": "Codex", "commit": "",
        "notes": "No activity, dispatch, build or fuel-share target. Detailed endogenous formulation retained only as rejected runtime evidence.",
    }])
    snapshot = ledger / "snapshots" / "biomass_simple_v23_2026-08-24.json"
    write_json(snapshot, {
        "schema": "philippines-v23-biomass-simple-source-extract-v1",
        "classifications": {"cooking service": "exogenous final demand", "legacy cooking and CHP": "initial stocks", "generic and residue caps": "continuing physical upper constraints", "observed shares": "benchmark only", "activity and dispatch": "endogenous"},
        "values": {"generic_cap_pj": GENERIC_CAP_PJ, "residue_cap_pj": RESIDUE_CAP_PJ, "wood_components_pj": WOOD_COMPONENTS, "residue_components_pj": RESIDUE_COMPONENTS, "generic_cost_musd_pj": GENERIC_COST, "residue_cost_musd_pj": RESIDUE_COST, "charcoal_route_iar": CHARCOAL_ROUTE_IAR, "cooking_biomass_input_2020_pj": COOKING_BIOMASS_INPUT_2020_PJ, "cooking_useful_2020_pj": COOKING_USEFUL_2020_PJ, "cooking_scale": COOKING_SCALE, "capacity_to_activity": STANDARD_CAPACITY_TO_ACTIVITY, "generic_capacity_stock": nonbinding_capacity_stock(GENERIC_CAP_PJ), "residue_capacity_stock": nonbinding_capacity_stock(RESIDUE_CAP_PJ)},
        "deferred": "Crop/forest-responsive ceilings and separate generated/collected/uncollected material accounts require a later sparse formulation and runtime A/B.",
    })
    note = ledger / "calculation_notes" / "MODEL_FIXES_BIOMASS_SIMPLE_V23_2026-08-24.md"
    note.parent.mkdir(parents=True, exist_ok=True)
    note.write_text(f"""# Philippines v23 minimal biomass-supply correction

The undocumented founding `PHL_PRO_PROC_BIOM` constant is replaced by a documented finite generic-biomass supply boundary at {GENERIC_COST:.12g} MUSD/PJ and {GENERIC_CAP_PJ:.12g} PJ/year. One new technology provides at most {RESIDUE_CAP_PJ:.12g} PJ/year of crop residues at {RESIDUE_COST:.12g} MUSD/PJ; its two modes allocate the same resource between the shared legacy-CHP pool and generic biomass. Both legacy CHP slices consume the same commodity at IAR 4.000666667. No minimum activity or share is imposed.

DOE's {COOKING_BIOMASS_INPUT_2020_PJ:.12g} PJ household-biomass observation is multiplied by the normalized PSA cooking-household-share-weighted route efficiency ({COOKING_WEIGHTED_EFFICIENCY:.12g}), yielding a {COOKING_USEFUL_2020_PJ:.12g} PJ useful-cooking proxy. The inherited cooking demand and initial-stock envelopes therefore scale by {COOKING_SCALE:.12g}. This is not represented as an exact fuel-energy balance: the public tables report household biomass energy and primary cooking fuel on different statistical bases. A fuel-specific cooking-energy calibration remains a high-priority gap, and the 2020 cooking-PM2.5 benchmark is not improved by this proxy. The closed legacy charcoal route consumes {CHARCOAL_ROUTE_IAR:.12g} PJ raw biomass per PJ useful service, incorporating traditional-kiln loss, and carries kiln CO2e and PM2.5 factors. No future charcoal-stove activity is forced.

This formulation intentionally defers endogenous crop/forest coupling. Its flat 2020 caps do not respond when modeled land allocation changes, and it does not separately report generated, collected and uncollected residues. Those are recorded as high/medium-priority gaps. The earlier detailed formulation added seven commodities and eight technologies; after fixing a 136-row TCC contradiction and two near-dependent balance designs, its BASE still exceeded the accepted 1.5x runtime gate. It is rejected, not promoted.

The supply technologies use the model-standard `CapacityToActivityUnit` of {STANDARD_CAPACITY_TO_ACTIVITY:.12g}. Their accounting-only residual capacities are {nonbinding_capacity_stock(GENERIC_CAP_PJ):.12g} and {nonbinding_capacity_stock(RESIDUE_CAP_PJ):.12g}, calculated so `ResidualCapacity × CapacityToActivityUnit = annual ceiling ÷ minimum YearSplit`. This is algebraically identical to the prior non-binding rate envelope while removing its outlying 2142.704 coefficient. A proposed `5 × annual ceiling` rate bound was not used because it would shrink the feasible set using an unsourced intrayear delivery limit.

Required qualification is: four source-only gates; application generation and preprocessing; GLPK matrix check; one BASE CBC optimum; then all policy cases and result validation in the same clean staging case. Seal the complete candidate and promote it unchanged with `scripts/sealed_case_promotion.py`; promotion performs no generation, preprocessing, matrix build or solver run. Runtime ratios are reported, but the scenario-specific COAL_PHASEOUT increase is not judged against a fixed historical ratio because the corrected biomass formulation changes the active policy problem. Failed diagnostic runs must remain outside the sealed candidate.
""", encoding="utf-8")
    documentation = target / "documentation" / "MODEL_FIXES_BIOMASS_SIMPLE_V23_2026-08-24.md"
    documentation.write_text(note.read_text(encoding="utf-8"), encoding="utf-8")
    return snapshot, write_workbook(ledger)


def build(target: Path):
    from Classes.Base import Config
    from Classes.Case.UpdateCaseClass import UpdateCase
    if target.exists():
        raise FileExistsError(target)
    source_hashes = {path.name: sha256(path) for path in SOURCE.glob("*.json")}
    shutil.copytree(SOURCE, target, ignore=shutil.ignore_patterns("res", "data_sources", ".DS_Store"))
    gen = read_json(target / "genData.json")
    mutate_structure(gen)
    Config.DATA_STORAGE = STORAGE
    UpdateCase(target.name, gen).updateCase()
    write_json(target / "genData.json", gen)
    overlay_parameters(target)
    # UpdateCase rewrites some untouched JSON files cosmetically.  Restore the
    # parent's exact bytes whenever parsed content is identical so promotion
    # has a minimal, byte-auditable source allowlist.
    for source_path in SOURCE.glob("*.json"):
        candidate_path = target / source_path.name
        if candidate_path.is_file() and read_json(candidate_path) == read_json(source_path):
            shutil.copy2(source_path, candidate_path)
    snapshot, workbook = write_ledger(target, source_hashes)
    manifest = {
        "schema": "philippines-v23-biomass-simple-build-v1", "parent_case": str(SOURCE), "candidate_case": str(target),
        "canonical_baseline": str(STORAGE / ".Philippines_v23-package1-candidate-20260824"),
        "source_hashes": source_hashes, "candidate_hashes": {path.name: sha256(path) for path in target.glob("*.json")},
        "changed_source_files": sorted(path.name for path in target.glob("*.json") if source_hashes.get(path.name) != sha256(path)),
        "source_snapshot": str(snapshot), "schema_workbook": str(workbook), "optimizer_runs": 0, "model_generation_runs": 0,
        "required_next_step": "Run all zero-solve gates before application generation.",
    }
    write_json(target / "documentation" / "biomass_simple_v23_build_manifest.json", manifest)
    print(json.dumps(manifest, indent=2))


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--target", type=Path, default=DEFAULT_TARGET)
    args = parser.parse_args()
    build(args.target.resolve())


if __name__ == "__main__":
    main()
